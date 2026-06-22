"""
Read/write the Compound_Table_Schema_v1.xlsx format.

Preserves all non-data sheets (README with formulas, Data_Dictionary,
Systemicity_Rubric, Gate_Mapping, Controlled_Vocab) untouched.
"""

import copy
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd


COLUMN_NAMES = [
    "compound_id", "common_name", "cas_rn", "dtxsid", "pubchem_cid",
    "inchikey", "canonical_smiles",
    "compound_category", "chemical_class", "module7_class", "application_route",
    "systemicity_class", "systemicity_confidence", "mobility_direction",
    "measured_mobility_value", "measured_mobility_metric", "label_source",
    "logKow_value", "logKow_provenance", "logD_pH5p5", "mw",
    "water_sol_mgL", "water_sol_provenance", "melting_point_C",
    "vapor_pressure_mPa", "abraham_E", "abraham_S", "abraham_A",
    "abraham_B", "abraham_V",
    "pka_acidic", "pka_basic", "pka_provenance", "tpsa", "hbd", "hba",
    "rotatable_bonds", "net_charge_pH7p5", "net_charge_pH5p5",
    "known_labile_groups", "gate5_alerts_predicted",
    "metabolic_halflife_plant", "known_metabolites_smiles",
    "biotransformer_runID",
    "descriptor_provenance_flag", "data_split", "stratum_origin",
    "entry_notes", "curator", "entry_date",
]

REQUIRED_FIELDS = [
    "compound_id", "common_name", "cas_rn", "canonical_smiles",
    "compound_category", "module7_class", "systemicity_class",
    "systemicity_confidence", "label_source", "logKow_value",
    "logKow_provenance", "mw", "descriptor_provenance_flag",
    "data_split", "stratum_origin", "curator", "entry_date",
]

CONTROLLED_VOCABS = {
    "compound_category": [
        "herbicide", "fungicide", "insecticide", "acaricide", "nematicide",
        "PGR", "phytohormone", "safener", "natural_product",
        "pharmaceutical", "industrial_pollutant", "other",
    ],
    "module7_class": [
        "WEAK_ACID_ORGANIC", "WEAK_ACID_SESQUITERPENOID",
        "PURINE_CYTOKININ", "PHENYLUREA_CYTOKININ", "CARRIER_GA",
        "NEUTRAL_LIPOPHILE", "HIGH_LIPOPHILE", "ZWITTERION",
        "STEROID_HORMONE", "STRIGOLACTONE_CANONICAL", "IRIDOID_GLYCOSIDE",
        "TERTIARY_AMINE_ALKALOID", "ALKALOID_N_OXIDE",
        "PHENOLIC_FLAVONOID", "PHOSPHONATE_NUTRIENT_MIMIC",
        "POLAR_NEUTRAL_SYSTEMIC", "AMINO_ACID_NEUTRAL",
        "AMINO_ACID_ACIDIC", "AMINO_ACID_BASIC",
        "OTHER_permanent_dication",
    ],
    "systemicity_class": [
        "0_contact", "1_local_translaminar", "2_xylem_only",
        "3_phloem_ambimobile", "unknown",
    ],
    "systemicity_confidence": ["high", "medium", "low"],
    "mobility_direction": [
        "acropetal", "basipetal", "ambimobile", "none", "unknown",
    ],
    "measured_mobility_metric": [
        "Kleier_Cf", "phloem_xylem_ratio", "RCF", "TSCF",
        "phloem_conc_uM", "qualitative", "none",
    ],
    "logKow_provenance": [
        "experimental", "predicted_OPERA", "predicted_other", "SEED",
    ],
    "water_sol_provenance": [
        "experimental", "predicted", "predicted_other", "SEED",
    ],
    "pka_provenance": [
        "experimental", "predicted_OPERA", "predicted_other", "SEED",
    ],
    "descriptor_provenance_flag": [
        "all_experimental", "mixed", "all_predicted",
    ],
    "data_split": ["train", "validation", "test_heldout"],
    "stratum_origin": [
        "natural_product", "synthetic_agrochem", "pollutant", "pharma",
    ],
    "application_route": [
        "foliar", "soil", "seed", "endogenous/foliar", "soil/foliar/seed",
    ],
}

DATA_START_ROW = 3  # row 1 = group headers, row 2 = column names


def read_compound_table(xlsx_path):
    """Read the Compound_Data sheet into a pandas DataFrame."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Compound_Data"]

    header_row = [c.value for c in ws[2]]
    col_map = {}
    for idx, name in enumerate(header_row):
        if name:
            col_map[idx] = name

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=len(header_row)):
        values = [c.value for c in row]
        if not any(values):
            continue
        row_dict = {}
        for idx, name in col_map.items():
            row_dict[name] = values[idx] if idx < len(values) else None
        rows.append(row_dict)

    wb.close()
    if not rows:
        return pd.DataFrame(columns=COLUMN_NAMES)
    return pd.DataFrame(rows)


def write_compound_row(xlsx_path, row_dict):
    """Append a single compound row to the Compound_Data sheet.

    Preserves all other sheets (README formulas, etc.) untouched.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Compound_Data"]

    header_row = [c.value for c in ws[2]]
    name_to_col = {}
    for idx, name in enumerate(header_row):
        if name:
            name_to_col[name] = idx + 1  # openpyxl is 1-based

    next_row = ws.max_row + 1
    # Skip if the last row is completely empty (max_row can overcount)
    while next_row > DATA_START_ROW:
        has_data = any(ws.cell(row=next_row - 1, column=c).value
                       for c in name_to_col.values())
        if has_data:
            break
        next_row -= 1
    if next_row < DATA_START_ROW:
        next_row = DATA_START_ROW

    for name, col_idx in name_to_col.items():
        val = row_dict.get(name)
        if val is not None:
            ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(xlsx_path)
    wb.close()
    return next_row


def write_compound_table(xlsx_path, df):
    """Write a full DataFrame to the Compound_Data sheet.

    Clears existing data rows and writes the DataFrame, preserving
    header rows and all other sheets.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Compound_Data"]

    header_row = [c.value for c in ws[2]]
    name_to_col = {}
    for idx, name in enumerate(header_row):
        if name:
            name_to_col[name] = idx + 1

    # Clear existing data rows
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        for col_idx in range(1, len(header_row) + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    # Write new data
    for i, (_, row) in enumerate(df.iterrows()):
        row_idx = DATA_START_ROW + i
        for name, col_idx in name_to_col.items():
            val = row.get(name)
            if pd.notna(val):
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(xlsx_path)
    wb.close()
    return len(df)


def validate_row(row_dict):
    """Check required fields and enum values.

    Returns a list of warning strings (empty = valid).
    """
    warnings = []

    for field in REQUIRED_FIELDS:
        val = row_dict.get(field)
        if val is None or (isinstance(val, str) and not val.strip()):
            warnings.append(f"MISSING required field: {field}")

    for field, allowed in CONTROLLED_VOCABS.items():
        val = row_dict.get(field)
        if val is None or (isinstance(val, str) and not val.strip()):
            continue
        val_str = str(val).strip()
        # application_route allows compound values like "soil/foliar/seed"
        if field == "application_route":
            continue
        if val_str not in allowed:
            warnings.append(
                f"INVALID value for {field}: '{val_str}' "
                f"(allowed: {', '.join(allowed[:5])}...)"
            )

    # pKa required for ionizables
    smiles = row_dict.get("canonical_smiles", "")
    m7 = row_dict.get("module7_class", "")
    if m7 and "ACID" in m7 and not row_dict.get("pka_acidic"):
        warnings.append("MISSING pka_acidic for acid-class compound")
    if m7 and ("AMINE" in m7 or "ALKALOID" in m7) and not row_dict.get("pka_basic"):
        warnings.append("MISSING pka_basic for base-class compound")

    # Circularity guard
    split = row_dict.get("data_split", "")
    prov = row_dict.get("descriptor_provenance_flag", "")
    if split in ("validation", "test_heldout") and prov == "all_predicted":
        warnings.append(
            "CIRCULARITY: test/validation row uses all_predicted descriptors"
        )

    return warnings


def get_controlled_vocab(xlsx_path=None, field_name=None):
    """Return allowed values for an enum field.

    Uses the built-in dict if xlsx_path is None; otherwise reads from
    the Controlled_Vocab sheet for the given field.
    """
    if field_name and field_name in CONTROLLED_VOCABS:
        return list(CONTROLLED_VOCABS[field_name])

    if xlsx_path is None or field_name is None:
        return dict(CONTROLLED_VOCABS)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Controlled_Vocab"]

    result = []
    found_field = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and str(row[0]).strip() == field_name:
            found_field = True
            vals_str = str(row[1]).strip() if row[1] else ""
            result = [v.strip() for v in vals_str.split("|") if v.strip()]
            break

    wb.close()
    return result if found_field else CONTROLLED_VOCABS.get(field_name, [])


def get_label_balance(xlsx_path):
    """Return a dict of systemicity_class counts from the compound table."""
    df = read_compound_table(xlsx_path)
    if df.empty or "systemicity_class" not in df.columns:
        return {}

    counts = df["systemicity_class"].value_counts().to_dict()
    total = len(df)

    # Ensure all classes present
    for cls in CONTROLLED_VOCABS["systemicity_class"]:
        counts.setdefault(cls, 0)

    counts["_total"] = total
    non_phloem = sum(
        counts.get(c, 0) for c in
        ["0_contact", "1_local_translaminar", "2_xylem_only"]
    )
    counts["_negative_pct"] = (
        round(100.0 * non_phloem / total, 1) if total > 0 else 0.0
    )
    return counts


def next_compound_id(xlsx_path):
    """Generate the next PSB-NNNN compound ID."""
    df = read_compound_table(xlsx_path)
    if df.empty:
        return "PSB-0001"

    max_num = 0
    for cid in df["compound_id"].dropna():
        cid_str = str(cid)
        if cid_str.startswith("PSB-"):
            try:
                num = int(cid_str.split("-")[1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
    return f"PSB-{max_num + 1:04d}"
