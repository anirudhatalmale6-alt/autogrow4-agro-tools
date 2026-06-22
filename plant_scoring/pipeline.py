"""
Plant Systemic Bioavailability — orchestration pipeline.

Ties together standardize → descriptors → schema_io into a single
entry point for processing compounds.
"""

import argparse
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from rdkit import Chem

from plant_scoring.standardize import (
    standardize_smiles,
    fetch_identifiers,
    detect_ionizable_groups,
    resolve_protonation,
)
from plant_scoring.descriptors import (
    compute_descriptors,
    descriptor_provenance_flag,
    detect_labile_groups,
)
from plant_scoring.classifier import classify_compound
from plant_scoring.gates import score_all_gates
from plant_scoring import schema_io

try:
    from plant_scoring.gate5 import detect_alerts, compute_gate5_multiplier
    _HAS_GATE5 = True
except ImportError:
    _HAS_GATE5 = False


# The 8 seed compounds from Compound_Table_Schema_v1.xlsx
SEED_COMPOUNDS = [
    {
        "compound_id": "PSB-0001",
        "common_name": "Glyphosate",
        "smiles": "OC(=O)CNCP(=O)(O)O",
        "compound_category": "herbicide",
        "systemicity_class": "3_phloem_ambimobile",
        "stratum_origin": "synthetic_agrochem",
    },
    {
        "compound_id": "PSB-0002",
        "common_name": "2,4-D",
        "smiles": "OC(=O)COc1ccc(Cl)cc1Cl",
        "compound_category": "herbicide",
        "systemicity_class": "3_phloem_ambimobile",
        "stratum_origin": "synthetic_agrochem",
    },
    {
        "compound_id": "PSB-0003",
        "common_name": "Indole-3-acetic acid (IAA)",
        "smiles": "OC(=O)Cc1c[nH]c2ccccc12",
        "compound_category": "phytohormone",
        "systemicity_class": "3_phloem_ambimobile",
        "stratum_origin": "natural_product",
    },
    {
        "compound_id": "PSB-0004",
        "common_name": "Imidacloprid",
        "smiles": "O=[N+]([O-])/N=C1\\NCCN1Cc1ccc(Cl)nc1",
        "compound_category": "insecticide",
        "systemicity_class": "3_phloem_ambimobile",
        "stratum_origin": "synthetic_agrochem",
    },
    {
        "compound_id": "PSB-0005",
        "common_name": "Tebuconazole",
        "smiles": "CC(C)(C)C(O)(Cn1cncn1)CCc1ccc(Cl)cc1",
        "compound_category": "fungicide",
        "systemicity_class": "2_xylem_only",
        "stratum_origin": "synthetic_agrochem",
    },
    {
        "compound_id": "PSB-0006",
        "common_name": "Nicotine",
        "smiles": "CN1CCC[C@H]1c1cccnc1",
        "compound_category": "insecticide_botanical",
        "systemicity_class": "2_xylem_only",
        "stratum_origin": "natural_product",
    },
    {
        "compound_id": "PSB-0007",
        "common_name": "Paraquat dichloride",
        "smiles": "C[n+]1ccc(-c2cc[n+](C)cc2)cc1",
        "compound_category": "herbicide",
        "systemicity_class": "0_contact",
        "stratum_origin": "synthetic_agrochem",
    },
    {
        "compound_id": "PSB-0008",
        "common_name": "Chlorothalonil",
        "smiles": "N#Cc1c(Cl)c(C#N)c(Cl)c(Cl)c1Cl",
        "compound_category": "fungicide",
        "systemicity_class": "0_contact",
        "stratum_origin": "synthetic_agrochem",
    },
]


def process_compound(
    smiles,
    compound_id=None,
    common_name=None,
    compound_category=None,
    experimental_overrides=None,
    curator="pipeline_auto",
    use_pubchem=False,
    stratum_origin=None,
    systemicity_class=None,
    extra_fields=None,
):
    """Process a single SMILES through the full pipeline.

    Returns
    -------
    dict with keys:
      row : dict  — schema-compatible row ready for write_compound_row
      warnings : list[str]  — validation warnings
      mol : rdkit.Chem.Mol or None
      std_result : dict  — full standardization result
      descriptors : dict  — raw descriptor dict from compute_descriptors
    """
    result = {
        "row": None,
        "warnings": [],
        "mol": None,
        "std_result": None,
        "descriptors": None,
    }

    # 1. Standardize
    std = standardize_smiles(smiles)
    result["std_result"] = std

    if not std["is_valid"]:
        result["warnings"].append(
            f"INVALID SMILES: {smiles} — {std['standardization_log']}"
        )
        return result

    mol = std["mol"]
    result["mol"] = mol

    # 2. Optional PubChem lookup
    pubchem = None
    if use_pubchem:
        pubchem = fetch_identifiers(std["canonical_smiles"], use_pubchem=True)

    # 3. Compute all descriptors
    desc = compute_descriptors(mol, experimental_overrides)
    result["descriptors"] = desc

    # 4. Detect labile groups
    labile = detect_labile_groups(mol)

    # 5. Determine provenance flag
    prov_flag = descriptor_provenance_flag(desc)

    # 6. Net charges via protonation resolution
    prot_55 = resolve_protonation(mol, 5.5)
    prot_75 = resolve_protonation(mol, 7.5)

    # 7. Build the row dict
    row = {}

    # Identity
    row["compound_id"] = compound_id or (
        f"PSB-AUTO-{abs(hash(smiles)) % 100000:05d}"
    )
    row["common_name"] = common_name or (
        pubchem.get("common_name") if pubchem else None
    ) or ""
    row["cas_rn"] = (pubchem.get("cas_rn") if pubchem else None) or ""
    row["dtxsid"] = (pubchem.get("dtxsid") if pubchem else None) or ""
    row["pubchem_cid"] = (pubchem.get("cid") if pubchem else None) or ""
    row["inchikey"] = std.get("inchikey") or ""
    row["canonical_smiles"] = std["canonical_smiles"]

    # Classification via Module 7
    classification = classify_compound(mol, desc)
    result["classification"] = classification

    row["compound_category"] = compound_category or ""
    row["chemical_class"] = ""
    row["module7_class"] = classification["module7_class"]
    row["application_route"] = ""

    # Target (labels set by the PI, not the pipeline)
    row["systemicity_class"] = systemicity_class or ""
    row["systemicity_confidence"] = ""
    row["mobility_direction"] = ""
    row["measured_mobility_value"] = ""
    row["measured_mobility_metric"] = ""
    row["label_source"] = ""

    # Gate 1 descriptors
    for key in [
        "logKow_value", "logD_pH5p5", "mw", "water_sol_mgL",
        "melting_point_C", "vapor_pressure_mPa",
        "abraham_E", "abraham_S", "abraham_A", "abraham_B", "abraham_V",
    ]:
        d = desc.get(key, {})
        row[key] = d.get("value")

    # logKow provenance
    lk = desc.get("logKow_value", {})
    row["logKow_provenance"] = _map_provenance(lk.get("provenance"))

    # Water solubility provenance
    ws = desc.get("water_sol_mgL", {})
    row["water_sol_provenance"] = _map_provenance(ws.get("provenance"))

    # Gate 2/3 descriptors
    for key in [
        "pka_acidic", "pka_basic", "tpsa", "hbd", "hba",
        "rotatable_bonds", "net_charge_pH7p5", "net_charge_pH5p5",
    ]:
        d = desc.get(key, {})
        row[key] = d.get("value")

    # pKa provenance
    pk = desc.get("pka_acidic", {})
    row["pka_provenance"] = _map_provenance(pk.get("provenance"))

    # Gate 5 alerts
    gate5_mult = 1.0
    gate5_alert_ids = []
    if _HAS_GATE5:
        alerts = detect_alerts(mol)
        gate5_mult, g5_details = compute_gate5_multiplier(
            mol, alerts=alerts, classification=classification)
        gate5_alert_ids = g5_details.get("fired_alerts", [])
        result["gate5_details"] = g5_details

    row["known_labile_groups"] = ";".join(labile) if labile else ""
    row["gate5_alerts_predicted"] = ";".join(gate5_alert_ids) if gate5_alert_ids else ""
    row["metabolic_halflife_plant"] = ""
    row["known_metabolites_smiles"] = ""
    row["biotransformer_runID"] = ""

    # Score all gates
    scoring = score_all_gates(desc, classification,
                              gate5_alerts=gate5_alert_ids,
                              gate5_multiplier=gate5_mult)
    result["scoring"] = scoring
    comp = scoring["composite_result"]
    result["composite_score"] = comp["composite"]
    result["bottleneck"] = comp["bottleneck"]
    result["design_flags"] = comp.get("design_flags", [])

    # QA
    row["descriptor_provenance_flag"] = prov_flag
    row["data_split"] = ""
    row["stratum_origin"] = stratum_origin or ""
    row["entry_notes"] = ""
    row["curator"] = curator
    row["entry_date"] = str(date.today())

    # Merge any extra fields (for seed compound overrides)
    if extra_fields:
        for k, v in extra_fields.items():
            if v is not None and v != "":
                row[k] = v

    result["row"] = row

    # 8. Validate
    result["warnings"] = schema_io.validate_row(row)

    return result


def _map_provenance(raw_prov):
    """Map internal provenance strings to schema enum values."""
    if raw_prov is None:
        return ""
    prov = str(raw_prov).lower()
    if "experimental" in prov:
        return "experimental"
    if "opera" in prov:
        return "predicted_OPERA"
    if "seed" in prov:
        return "SEED"
    if "predicted" in prov or "rdkit" in prov or "smarts" in prov or "approx" in prov:
        return "predicted_other"
    return "predicted_other"


def process_batch(smiles_list, output_xlsx=None, **kwargs):
    """Process multiple SMILES and optionally write to xlsx.

    Parameters
    ----------
    smiles_list : list of str or list of dict
        If str, just SMILES.  If dict, must have "smiles" key and
        optionally compound_id, common_name, compound_category, etc.
    output_xlsx : str or Path, optional
        If given, writes results to this xlsx file (creates or appends).

    Returns
    -------
    list of dict — one per input, each with row/warnings/mol keys.
    """
    results = []
    for item in smiles_list:
        if isinstance(item, str):
            smi = item
            extra = {}
        elif isinstance(item, dict):
            smi = item["smiles"]
            extra = {k: v for k, v in item.items() if k != "smiles"}
        else:
            results.append({
                "row": None,
                "warnings": [f"Unrecognised input type: {type(item)}"],
                "mol": None,
            })
            continue

        merged_kwargs = dict(kwargs)
        for key in [
            "compound_id", "common_name", "compound_category",
            "stratum_origin", "systemicity_class",
        ]:
            if key in extra and key not in merged_kwargs:
                merged_kwargs[key] = extra.pop(key)
        if extra:
            merged_kwargs.setdefault("extra_fields", {}).update(extra)

        r = process_compound(smi, **merged_kwargs)
        results.append(r)

    if output_xlsx:
        output_path = Path(output_xlsx)
        rows = [r["row"] for r in results if r["row"] is not None]
        if rows:
            df = pd.DataFrame(rows)
            if output_path.exists():
                schema_io.write_compound_table(str(output_path), df)
            else:
                _create_minimal_xlsx(str(output_path), df)

    return results


def _create_minimal_xlsx(path, df):
    """Create a minimal xlsx with just headers and data (no fancy schema)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compound_Data"

    # Row 1: group headers (simplified)
    ws.cell(row=1, column=1, value="IDENTITY")
    ws.cell(row=1, column=8, value="CLASSIFICATION")
    ws.cell(row=1, column=12, value="TARGET")
    ws.cell(row=1, column=18, value="GATE 1")
    ws.cell(row=1, column=31, value="GATE 2/3")
    ws.cell(row=1, column=40, value="GATE 5")
    ws.cell(row=1, column=45, value="QA")

    # Row 2: column names
    for i, col_name in enumerate(schema_io.COLUMN_NAMES):
        ws.cell(row=2, column=i + 1, value=col_name)

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, col_name in enumerate(schema_io.COLUMN_NAMES):
            val = row.get(col_name)
            if pd.notna(val) and val != "":
                ws.cell(row=3 + row_idx, column=col_idx + 1, value=val)

    wb.save(path)
    wb.close()


def run_demo():
    """Process the 8 seed compounds and print a formatted report."""
    print("=" * 100)
    print("PLANT SYSTEMIC BIOAVAILABILITY PIPELINE — DEMO")
    print(f"  Processing {len(SEED_COMPOUNDS)} seed compounds")
    print(f"  Date: {date.today()}")
    print("=" * 100)

    results = []
    for seed in SEED_COMPOUNDS:
        r = process_compound(
            smiles=seed["smiles"],
            compound_id=seed["compound_id"],
            common_name=seed["common_name"],
            compound_category=seed["compound_category"],
            stratum_origin=seed["stratum_origin"],
            systemicity_class=seed.get("systemicity_class", ""),
            use_pubchem=False,
            curator="demo",
        )
        results.append(r)

    # Table 1: Identity + basic descriptors
    print("\nTABLE 1: IDENTITY & BASIC DESCRIPTORS")
    print("=" * 100)
    print(f"{'ID':<10} {'Name':<22} {'SMILES':<35} {'InChIKey':<15}")
    print("-" * 100)
    for r in results:
        row = r["row"]
        if row is None:
            continue
        ik = (row.get("inchikey") or "")[:14]
        smi = (row.get("canonical_smiles") or "")[:34]
        print(f"{row['compound_id']:<10} {row['common_name']:<22} {smi:<35} {ik:<15}")

    # Table 2: Physico-chemical descriptors
    print(f"\nTABLE 2: PHYSICO-CHEMICAL DESCRIPTORS")
    print("=" * 100)
    print(f"{'ID':<10} {'logKow':>7} {'logD55':>7} {'MW':>7} {'Sol':>8} "
          f"{'pKa_a':>6} {'pKa_b':>6} {'TPSA':>6} {'HBD':>4} {'HBA':>4} "
          f"{'Q75':>4} {'Q55':>4}")
    print("-" * 100)
    for r in results:
        row = r["row"]
        if row is None:
            continue
        def _f(v, w=7, d=2):
            if v is None or v == "":
                return "---".rjust(w)
            return f"{float(v):{w}.{d}f}"
        def _i(v, w=4):
            if v is None or v == "":
                return "---".rjust(w)
            return f"{int(round(float(v))):{w}d}"
        print(f"{row['compound_id']:<10} {_f(row.get('logKow_value'))} "
              f"{_f(row.get('logD_pH5p5'))} {_f(row.get('mw'))} "
              f"{_f(row.get('water_sol_mgL'), 8, 1)} "
              f"{_f(row.get('pka_acidic'), 6)} {_f(row.get('pka_basic'), 6)} "
              f"{_f(row.get('tpsa'), 6)} {_i(row.get('hbd'))} "
              f"{_i(row.get('hba'))} {_i(row.get('net_charge_pH7p5'))} "
              f"{_i(row.get('net_charge_pH5p5'))}")

    # Table 3: Labile groups + provenance
    print(f"\nTABLE 3: LABILE GROUPS & PROVENANCE")
    print("=" * 100)
    print(f"{'ID':<10} {'Labile Groups':<40} {'Provenance':<18} {'Origin':<20}")
    print("-" * 100)
    for r in results:
        row = r["row"]
        if row is None:
            continue
        lg = row.get("known_labile_groups") or "---"
        prov = row.get("descriptor_provenance_flag") or "---"
        origin = row.get("stratum_origin") or "---"
        print(f"{row['compound_id']:<10} {lg:<40} {prov:<18} {origin:<20}")

    # Table 4: Abraham descriptors
    print(f"\nTABLE 4: ABRAHAM SOLUTE DESCRIPTORS (pp-LFER, approximate)")
    print("=" * 100)
    print(f"{'ID':<10} {'E':>8} {'S':>8} {'A':>8} {'B':>8} {'V':>8}")
    print("-" * 100)
    for r in results:
        row = r["row"]
        if row is None:
            continue
        def _fa(v, w=8):
            if v is None or v == "":
                return "---".rjust(w)
            return f"{float(v):{w}.3f}"
        print(f"{row['compound_id']:<10} {_fa(row.get('abraham_E'))} "
              f"{_fa(row.get('abraham_S'))} {_fa(row.get('abraham_A'))} "
              f"{_fa(row.get('abraham_B'))} {_fa(row.get('abraham_V'))}")

    # Table 5: Gate Scores + Composite
    print(f"\nTABLE 5: GATE SCORES & COMPOSITE")
    print("=" * 120)
    print(f"{'ID':<10} {'Class':<28} {'G1':>5} {'G2':>5} {'G3':>5} "
          f"{'G4':>5} {'G5':>5} {'Comp':>6} {'Bottleneck':<12} {'Flags'}")
    print("-" * 120)
    for r in results:
        row = r.get("row")
        if row is None:
            continue
        sc = r.get("scoring", {})
        comp = sc.get("composite_result", {})
        gs = comp.get("gate_scores", {})
        flags = ", ".join(comp.get("design_flags", [])[:3]) or "---"
        cls_name = row.get("module7_class", "---")
        if len(cls_name) > 27:
            cls_name = cls_name[:25] + ".."
        print(f"{row['compound_id']:<10} {cls_name:<28} "
              f"{gs.get('gate1', 0):5.3f} {gs.get('gate2', 0):5.3f} "
              f"{gs.get('gate3', 0):5.3f} {gs.get('gate4', 0):5.3f} "
              f"{gs.get('gate5', 1):5.3f} {comp.get('composite', 0):6.4f} "
              f"{comp.get('bottleneck', '---'):<12} {flags}")

    # Validation summary
    print(f"\nVALIDATION WARNINGS")
    print("=" * 100)
    any_warn = False
    for r in results:
        row = r.get("row")
        cid = row["compound_id"] if row else "???"
        for w in r.get("warnings", []):
            print(f"  {cid}: {w}")
            any_warn = True
    if not any_warn:
        print("  No warnings.")

    # Label balance
    print(f"\nLABEL BALANCE (seed set)")
    print("-" * 50)
    counts = {}
    for r in results:
        row = r.get("row")
        if row:
            cls = row.get("systemicity_class", "unknown")
            counts[cls] = counts.get(cls, 0) + 1
    total = sum(counts.values())
    for cls in schema_io.CONTROLLED_VOCABS["systemicity_class"]:
        c = counts.get(cls, 0)
        pct = 100.0 * c / total if total else 0
        print(f"  {cls:<30} {c:>3}  ({pct:5.1f}%)")
    neg = sum(counts.get(c, 0) for c in
              ["0_contact", "1_local_translaminar", "2_xylem_only"])
    neg_pct = 100.0 * neg / total if total else 0
    print(f"  {'NEGATIVES (0+1+2)':<30} {neg:>3}  ({neg_pct:5.1f}%)")
    if neg_pct < 25:
        print("  ** WARNING: negatives < 25% — survivorship bias risk **")

    print(f"\nDone. Processed {len(results)} compounds.")
    return results


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Plant Systemic Bioavailability Pipeline"
    )
    parser.add_argument("--smiles", help="Single SMILES to process")
    parser.add_argument("--input", help="Input file (one SMILES per line, or .smi)")
    parser.add_argument("--output", help="Output xlsx path")
    parser.add_argument("--demo", action="store_true", help="Run demo on seed compounds")
    parser.add_argument("--pubchem", action="store_true",
                        help="Enable PubChem lookups for identifiers")
    parser.add_argument("--curator", default="pipeline_auto",
                        help="Curator name for entry_notes")
    args = parser.parse_args()

    if args.demo:
        results = run_demo()
        if args.output and results:
            rows = [r["row"] for r in results if r["row"] is not None]
            if rows:
                df = pd.DataFrame(rows)
                _create_minimal_xlsx(args.output, df)
                print(f"\nSaved to {args.output}")
        return

    if args.smiles:
        r = process_compound(
            args.smiles,
            use_pubchem=args.pubchem,
            curator=args.curator,
        )
        if r["row"]:
            print(f"\nCompound: {r['row']['canonical_smiles']}")
            print(f"  InChIKey:  {r['row'].get('inchikey', '---')}")
            print(f"  MW:        {r['row'].get('mw', '---')}")
            print(f"  logKow:    {r['row'].get('logKow_value', '---')}")
            print(f"  pKa acid:  {r['row'].get('pka_acidic', '---')}")
            print(f"  pKa base:  {r['row'].get('pka_basic', '---')}")
            print(f"  TPSA:      {r['row'].get('tpsa', '---')}")
            print(f"  Sol mg/L:  {r['row'].get('water_sol_mgL', '---')}")
            print(f"  Labile:    {r['row'].get('known_labile_groups', '---')}")
            print(f"  Prov:      {r['row'].get('descriptor_provenance_flag', '---')}")
            if r["warnings"]:
                print(f"\n  Warnings:")
                for w in r["warnings"]:
                    print(f"    - {w}")
        else:
            print(f"FAILED: {r['warnings']}")
        if args.output and r["row"]:
            out = Path(args.output)
            if out.exists():
                schema_io.write_compound_row(str(out), r["row"])
                print(f"\nAppended to {out}")
            else:
                _create_minimal_xlsx(str(out), pd.DataFrame([r["row"]]))
                print(f"\nCreated {out}")
        return

    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} not found")
            sys.exit(1)

        smiles_list = []
        with open(input_path) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split()
                smi = parts[0]
                name = parts[1] if len(parts) > 1 else None
                item = {"smiles": smi}
                if name:
                    item["common_name"] = name
                smiles_list.append(item)

        if not smiles_list:
            print("No SMILES found in input file")
            sys.exit(1)

        results = process_batch(
            smiles_list,
            output_xlsx=args.output,
            use_pubchem=args.pubchem,
            curator=args.curator,
        )

        ok = sum(1 for r in results if r["row"] is not None)
        fail = len(results) - ok
        print(f"\nProcessed {len(results)} compounds: {ok} OK, {fail} failed")
        if args.output:
            print(f"Output: {args.output}")
        return

    parser.print_help()


if __name__ == "__main__":
    main()
